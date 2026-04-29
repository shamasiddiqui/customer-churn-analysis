




import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import random

np.random.seed(42)
random.seed(42)
n = 1000

# --- Generate Dataset ---
customer_ids = [f"CUST{str(i).zfill(4)}" for i in range(1, n+1)]
genders = np.random.choice(["Male", "Female"], n)
ages = np.random.randint(18, 70, n)
tenure = np.random.randint(1, 72, n)
monthly_charges = np.round(np.random.uniform(20, 120, n), 2)
total_charges = np.round(monthly_charges * tenure * np.random.uniform(0.9, 1.1, n), 2)
contract = np.random.choice(["Month-to-month", "One year", "Two year"], n, p=[0.55, 0.25, 0.20])
internet_service = np.random.choice(["DSL", "Fiber optic", "No"], n, p=[0.35, 0.45, 0.20])
payment_method = np.random.choice(["Electronic check", "Mailed check", "Bank transfer", "Credit card"], n)
tech_support = np.random.choice(["Yes", "No"], n)
online_security = np.random.choice(["Yes", "No"], n)
num_complaints = np.random.randint(0, 6, n)
satisfaction_score = np.random.randint(1, 6, n)

# Churn logic (realistic)
churn_prob = (
    0.05
    + 0.25 * (contract == "Month-to-month")
    + 0.10 * (internet_service == "Fiber optic")
    + 0.08 * (tech_support == "No")
    + 0.06 * (online_security == "No")
    + 0.04 * (num_complaints > 2)
    + 0.06 * (satisfaction_score <= 2)
    - 0.10 * (tenure > 36)
    - 0.05 * (satisfaction_score >= 4)
)
churn_prob = np.clip(churn_prob, 0.02, 0.85)
churn = np.where(np.random.rand(n) < churn_prob, "Yes", "No")

df = pd.DataFrame({
    "CustomerID": customer_ids,
    "Gender": genders,
    "Age": ages,
    "Tenure_Months": tenure,
    "Contract_Type": contract,
    "Internet_Service": internet_service,
    "Tech_Support": tech_support,
    "Online_Security": online_security,
    "Payment_Method": payment_method,
    "Monthly_Charges": monthly_charges,
    "Total_Charges": total_charges,
    "Num_Complaints": num_complaints,
    "Satisfaction_Score": satisfaction_score,
    "Churn": churn
})

# --- Summary Stats ---
churn_rate = round((df["Churn"] == "Yes").mean() * 100, 1)
avg_tenure_churn = round(df[df["Churn"]=="Yes"]["Tenure_Months"].mean(), 1)
avg_tenure_no = round(df[df["Churn"]=="No"]["Tenure_Months"].mean(), 1)
avg_charge_churn = round(df[df["Churn"]=="Yes"]["Monthly_Charges"].mean(), 2)
avg_charge_no = round(df[df["Churn"]=="No"]["Monthly_Charges"].mean(), 2)

contract_churn = df.groupby("Contract_Type")["Churn"].apply(lambda x: round((x=="Yes").mean()*100,1)).reset_index()
contract_churn.columns = ["Contract_Type", "Churn_Rate_%"]

internet_churn = df.groupby("Internet_Service")["Churn"].apply(lambda x: round((x=="Yes").mean()*100,1)).reset_index()
internet_churn.columns = ["Internet_Service", "Churn_Rate_%"]

segment_summary = df.groupby("Churn").agg(
    Count=("CustomerID","count"),
    Avg_Tenure=("Tenure_Months","mean"),
    Avg_Monthly_Charge=("Monthly_Charges","mean"),
    Avg_Satisfaction=("Satisfaction_Score","mean"),
    Avg_Complaints=("Num_Complaints","mean")
).round(2).reset_index()

# --- Build Excel Workbook ---
wb = Workbook()

# Colors
DARK_BLUE = "1F3864"
MED_BLUE  = "2E75B6"
LIGHT_BLUE= "D6E4F0"
ACCENT    = "E74C3C"
ACCENT2   = "27AE60"
WHITE     = "FFFFFF"
GRAY_HDR  = "F2F2F2"
ORANGE    = "F39C12"

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

thin = Side(style="thin", color="CCCCCC")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, col_start, col_end, bg=MED_BLUE, font_color=WHITE, height=28):
    for c in range(col_start, col_end+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = Font(name="Arial", size=10, bold=True, color=font_color)
        cell.alignment = center()
        cell.border = thin_border
    ws.row_dimensions[row].height = height

def style_data_row(ws, row, col_start, col_end, alt=False):
    bg = "EBF5FB" if alt else WHITE
    for c in range(col_start, col_end+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(bg)
        cell.font = body_font()
        cell.alignment = center()
        cell.border = thin_border

# ===================== SHEET 1: RAW DATA =====================
ws1 = wb.active
ws1.title = "Customer Data"
ws1.freeze_panes = "A2"

headers = list(df.columns)
ws1.append(headers)
style_header_row(ws1, 1, 1, len(headers), bg=DARK_BLUE)

for i, row in enumerate(df.itertuples(index=False), start=2):
    ws1.append(list(row))
    alt = (i % 2 == 0)
    for c in range(1, len(headers)+1):
        cell = ws1.cell(row=i, column=c)
        cell.fill = fill("EBF5FB" if alt else WHITE)
        cell.font = body_font()
        cell.alignment = center()
        cell.border = thin_border
        if headers[c-1] == "Churn":
            cell.font = Font(name="Arial", size=10, bold=True,
                             color=ACCENT if cell.value=="Yes" else "27AE60")

col_widths = [12,8,6,14,16,16,12,16,18,15,13,14,17,8]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title block
ws1.insert_rows(1)
ws1.merge_cells("A1:N1")
title_cell = ws1["A1"]
title_cell.value = "Customer Churn Dataset — Telecom Analytics Project"
title_cell.font = Font(name="Arial", size=13, bold=True, color=WHITE)
title_cell.fill = fill(DARK_BLUE)
title_cell.alignment = center()
ws1.row_dimensions[1].height = 32

# ===================== SHEET 2: SUMMARY DASHBOARD =====================
ws2 = wb.create_sheet("Summary Dashboard")
ws2.sheet_view.showGridLines = False

def merged_title(ws, cell, text, bg=DARK_BLUE, fsize=14):
    c = ws[cell]
    c.value = text
    c.font = Font(name="Arial", size=fsize, bold=True, color=WHITE)
    c.fill = fill(bg)
    c.alignment = center()

# Main title
ws2.merge_cells("A1:J1")
merged_title(ws2, "A1", "Customer Churn Analysis — Executive Dashboard", fsize=14)
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A2:J2")
sub = ws2["A2"]
sub.value = "Telecom Company | 1,000 Customers Analyzed"
sub.font = Font(name="Arial", size=10, italic=True, color="555555")
sub.alignment = center()
sub.fill = fill("F0F4F8")

# KPI Cards Row
kpis = [
    ("Total Customers", "1,000", MED_BLUE),
    ("Churned", f"{(df['Churn']=='Yes').sum()}", ACCENT),
    ("Churn Rate", f"{churn_rate}%", ACCENT),
    ("Avg Tenure (Churned)", f"{avg_tenure_churn} mo", ORANGE),
    ("Avg Tenure (Retained)", f"{avg_tenure_no} mo", ACCENT2),
]

kpi_cols = [1, 3, 5, 7, 9]
ws2.merge_cells("A4:J4")
ws2["A4"].value = "KEY PERFORMANCE INDICATORS"
ws2["A4"].font = Font(name="Arial", size=9, bold=True, color="888888")
ws2["A4"].alignment = left()
ws2.row_dimensions[4].height = 18

for idx, (label, value, color) in enumerate(kpis):
    col = kpi_cols[idx]
    end_col = col + 1
    ws2.merge_cells(start_row=5, start_column=col, end_row=5, end_column=end_col)
    ws2.merge_cells(start_row=6, start_column=col, end_row=6, end_column=end_col)
    ws2.merge_cells(start_row=7, start_column=col, end_row=7, end_column=end_col)

    top = ws2.cell(5, col)
    top.fill = fill(color)
    top.value = label
    top.font = Font(name="Arial", size=9, bold=True, color=WHITE)
    top.alignment = center()

    mid = ws2.cell(6, col)
    mid.fill = fill(color)
    mid.value = value
    mid.font = Font(name="Arial", size=18, bold=True, color=WHITE)
    mid.alignment = center()

    bot = ws2.cell(7, col)
    bot.fill = fill(color)
    bot.alignment = center()

    for r in [5,6,7]:
        for c in [col, col+1]:
            ws2.cell(r, c).fill = fill(color)
    ws2.row_dimensions[5].height = 20
    ws2.row_dimensions[6].height = 36
    ws2.row_dimensions[7].height = 14

# Spacer
ws2.row_dimensions[8].height = 10

# --- Churn by Contract ---
ws2["A9"].value = "CHURN RATE BY CONTRACT TYPE"
ws2["A9"].font = Font(name="Arial", size=9, bold=True, color="888888")
ws2.row_dimensions[9].height = 18

contract_headers = ["Contract Type", "Total Customers", "Churned", "Churn Rate %"]
for c_idx, h in enumerate(contract_headers, 1):
    cell = ws2.cell(10, c_idx)
    cell.value = h
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = fill(MED_BLUE)
    cell.alignment = center()
    cell.border = thin_border
ws2.row_dimensions[10].height = 24

contract_detail = df.groupby("Contract_Type").agg(
    Total=("CustomerID","count"),
    Churned=("Churn", lambda x: (x=="Yes").sum())
).reset_index()
contract_detail["Rate"] = (contract_detail["Churned"]/contract_detail["Total"]*100).round(1)

for r_idx, row in enumerate(contract_detail.itertuples(), start=11):
    alt = r_idx % 2 == 0
    bg = "EBF5FB" if alt else WHITE
    data = [row.Contract_Type, row.Total, row.Churned, f"{row.Rate}%"]
    for c_idx, val in enumerate(data, 1):
        cell = ws2.cell(r_idx, c_idx)
        cell.value = val
        cell.fill = fill(bg)
        cell.font = body_font(bold=(c_idx==4))
        if c_idx == 4:
            cell.font = Font(name="Arial", size=10, bold=True,
                             color=ACCENT if row.Rate > 30 else ACCENT2)
        cell.alignment = center()
        cell.border = thin_border
    ws2.row_dimensions[r_idx].height = 22

# --- Churn by Internet Service ---
ws2["F9"].value = "CHURN RATE BY INTERNET SERVICE"
ws2["F9"].font = Font(name="Arial", size=9, bold=True, color="888888")

inet_headers = ["Internet Service", "Total", "Churned", "Churn Rate %"]
for c_idx, h in enumerate(inet_headers, 6):
    cell = ws2.cell(10, c_idx)
    cell.value = h
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = fill(MED_BLUE)
    cell.alignment = center()
    cell.border = thin_border

inet_detail = df.groupby("Internet_Service").agg(
    Total=("CustomerID","count"),
    Churned=("Churn", lambda x: (x=="Yes").sum())
).reset_index()
inet_detail["Rate"] = (inet_detail["Churned"]/inet_detail["Total"]*100).round(1)

for r_idx, row in enumerate(inet_detail.itertuples(), start=11):
    alt = r_idx % 2 == 0
    bg = "EBF5FB" if alt else WHITE
    data = [row.Internet_Service, row.Total, row.Churned, f"{row.Rate}%"]
    for c_idx, val in enumerate(data, 6):
        cell = ws2.cell(r_idx, c_idx)
        cell.value = val
        cell.fill = fill(bg)
        cell.font = body_font()
        if c_idx == 9:
            cell.font = Font(name="Arial", size=10, bold=True,
                             color=ACCENT if row.Rate > 30 else ACCENT2)
        cell.alignment = center()
        cell.border = thin_border
    ws2.row_dimensions[r_idx].height = 22

# --- Segment Comparison ---
ws2.row_dimensions[14].height = 10
ws2["A15"].value = "CHURNED VS RETAINED — SEGMENT COMPARISON"
ws2["A15"].font = Font(name="Arial", size=9, bold=True, color="888888")
ws2.row_dimensions[15].height = 18

seg_headers = ["Segment", "Count", "Avg Tenure (mo)", "Avg Monthly Charge ($)", "Avg Satisfaction", "Avg Complaints"]
for c_idx, h in enumerate(seg_headers, 1):
    cell = ws2.cell(16, c_idx)
    cell.value = h
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = fill(DARK_BLUE)
    cell.alignment = center()
    cell.border = thin_border
ws2.row_dimensions[16].height = 28

for r_idx, row in enumerate(segment_summary.itertuples(), start=17):
    is_churn = row.Churn == "Yes"
    bg = "FEF0EF" if is_churn else "EAFAF1"
    label_color = ACCENT if is_churn else ACCENT2
    data = [
        "Churned" if is_churn else "Retained",
        row.Count,
        round(row.Avg_Tenure, 1),
        f"${round(row.Avg_Monthly_Charge, 2)}",
        round(row.Avg_Satisfaction, 2),
        round(row.Avg_Complaints, 2)
    ]
    for c_idx, val in enumerate(data, 1):
        cell = ws2.cell(r_idx, c_idx)
        cell.value = val
        cell.fill = fill(bg)
        cell.font = Font(name="Arial", size=10, bold=(c_idx==1), color=label_color if c_idx==1 else "000000")
        cell.alignment = center()
        cell.border = thin_border
    ws2.row_dimensions[r_idx].height = 22

# --- Key Insights ---
ws2.row_dimensions[20].height = 10
ws2.merge_cells("A21:J21")
ws2["A21"].value = "KEY INSIGHTS & RECOMMENDATIONS"
ws2["A21"].font = Font(name="Arial", size=9, bold=True, color="888888")
ws2.row_dimensions[21].height = 18

insights = [
    ("⚠ High Risk", f"Month-to-month customers churn at {contract_detail[contract_detail.Contract_Type=='Month-to-month']['Rate'].values[0]}% — offer annual discounts to convert them.", ACCENT),
    ("⚠ Service Issue", f"Fiber optic users churn more — investigate service quality and pricing.", ORANGE),
    ("✓ Loyalty Signal", f"Customers with 3+ years tenure churn {round(df[df.Tenure_Months>36]['Churn'].eq('Yes').mean()*100,1)}% — reward long-tenure customers.", ACCENT2),
    ("✓ Action", "Low satisfaction scores (1-2) are strong churn predictors — prioritize support outreach.", MED_BLUE),
]

for r_offset, (tag, text, color) in enumerate(insights, start=22):
    ws2.merge_cells(start_row=r_offset, start_column=1, end_row=r_offset, end_column=2)
    ws2.merge_cells(start_row=r_offset, start_column=3, end_row=r_offset, end_column=10)
    tag_cell = ws2.cell(r_offset, 1)
    tag_cell.value = tag
    tag_cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    tag_cell.fill = fill(color)
    tag_cell.alignment = center()
    tag_cell.border = thin_border

    txt_cell = ws2.cell(r_offset, 3)
    txt_cell.value = text
    txt_cell.font = Font(name="Arial", size=10, color="222222")
    txt_cell.fill = fill("F9F9F9")
    txt_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    txt_cell.border = thin_border
    ws2.row_dimensions[r_offset].height = 30

col_widths_2 = [18, 14, 16, 22, 22, 16, 16, 14, 10, 10]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ===================== SHEET 3: SQL QUERIES =====================
ws3 = wb.create_sheet("SQL Queries")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:H1")
ws3["A1"].value = "SQL Practice Queries — Customer Churn Analysis"
ws3["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws3["A1"].fill = fill(DARK_BLUE)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 32

sql_blocks = [
    ("Query 1: Overall churn rate",
     """SELECT
    COUNT(*) AS total_customers,
    SUM(CASE WHEN Churn = 'Yes' THEN 1 ELSE 0 END) AS churned,
    ROUND(
        100.0 * SUM(CASE WHEN Churn = 'Yes' THEN 1 ELSE 0 END) / COUNT(*), 1
    ) AS churn_rate_pct
FROM customers;"""),
    ("Query 2: Churn rate by contract type",
     """SELECT
    Contract_Type,
    COUNT(*) AS total,
    SUM(CASE WHEN Churn = 'Yes' THEN 1 ELSE 0 END) AS churned,
    ROUND(
        100.0 * SUM(CASE WHEN Churn = 'Yes' THEN 1 ELSE 0 END) / COUNT(*), 1
    ) AS churn_rate_pct
FROM customers
GROUP BY Contract_Type
ORDER BY churn_rate_pct DESC;"""),
    ("Query 3: Average metrics by churn status",
     """SELECT
    Churn,
    ROUND(AVG(Tenure_Months), 1)    AS avg_tenure,
    ROUND(AVG(Monthly_Charges), 2)  AS avg_monthly_charge,
    ROUND(AVG(Satisfaction_Score), 2) AS avg_satisfaction,
    ROUND(AVG(Num_Complaints), 2)   AS avg_complaints
FROM customers
GROUP BY Churn;"""),
    ("Query 4: High-risk customers (likely to churn)",
     """SELECT
    CustomerID, Contract_Type, Tenure_Months,
    Monthly_Charges, Satisfaction_Score, Num_Complaints
FROM customers
WHERE
    Contract_Type = 'Month-to-month'
    AND Satisfaction_Score <= 2
    AND Num_Complaints >= 2
    AND Churn = 'No'
ORDER BY Satisfaction_Score ASC, Num_Complaints DESC
LIMIT 20;"""),
    ("Query 5: Churn by internet service",
     """SELECT
    Internet_Service,
    COUNT(*) AS total,
    SUM(CASE WHEN Churn = 'Yes' THEN 1 ELSE 0 END) AS churned,
    ROUND(AVG(Monthly_Charges), 2) AS avg_charge,
    ROUND(
        100.0 * SUM(CASE WHEN Churn = 'Yes' THEN 1 ELSE 0 END) / COUNT(*), 1
    ) AS churn_rate_pct
FROM customers
GROUP BY Internet_Service
ORDER BY churn_rate_pct DESC;"""),
]

row_cursor = 3
for title, sql in sql_blocks:
    # Title
    ws3.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
    tc = ws3.cell(row_cursor, 1)
    tc.value = title
    tc.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    tc.fill = fill(MED_BLUE)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    tc.border = thin_border
    ws3.row_dimensions[row_cursor].height = 24
    row_cursor += 1

    for line in sql.split("\n"):
        ws3.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
        cc = ws3.cell(row_cursor, 1)
        cc.value = line
        cc.font = Font(name="Courier New", size=10, color="1A1A2E")
        cc.fill = fill("F4F6F9")
        cc.alignment = Alignment(horizontal="left", vertical="center")
        cc.border = thin_border
        ws3.row_dimensions[row_cursor].height = 18
        row_cursor += 1

    row_cursor += 1  # spacer

ws3.column_dimensions["A"].width = 100

# ===================== SHEET 4: PYTHON GUIDE =====================
ws4 = wb.create_sheet("Python Guide")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:H1")
ws4["A1"].value = "Python Analysis Code — Step by Step (Beginner Friendly)"
ws4["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws4["A1"].fill = fill(DARK_BLUE)
ws4["A1"].alignment = center()
ws4.row_dimensions[1].height = 32

py_blocks = [
    ("Step 1: Import libraries and load data",
     """import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

# Load the dataset (save the Customer Data sheet as CSV first)
df = pd.read_csv('customer_churn_data.csv')
print(df.shape)       # How many rows & columns?
print(df.head())      # Preview first 5 rows
print(df.info())      # Data types and missing values"""),
    ("Step 2: Basic churn rate",
     """# Count churned vs retained
churn_counts = df['Churn'].value_counts()
print(churn_counts)

# Churn rate as percentage
churn_rate = (df['Churn'] == 'Yes').mean() * 100
print(f'Churn Rate: {churn_rate:.1f}%')"""),
    ("Step 3: Churn by contract type (bar chart)",
     """# Group by contract type
contract_churn = df.groupby('Contract_Type')['Churn'].apply(
    lambda x: (x == 'Yes').mean() * 100
).reset_index()
contract_churn.columns = ['Contract_Type', 'Churn_Rate']

# Plot
plt.figure(figsize=(8, 5))
sns.barplot(data=contract_churn, x='Contract_Type', y='Churn_Rate', palette='Reds_r')
plt.title('Churn Rate by Contract Type', fontsize=14, fontweight='bold')
plt.ylabel('Churn Rate (%)')
plt.xlabel('Contract Type')
plt.tight_layout()
plt.savefig('churn_by_contract.png', dpi=150)
plt.show()"""),
    ("Step 4: Predict churn with logistic regression",
     """from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import classification_report, confusion_matrix

# Encode categorical columns
df_model = df.copy()
le = LabelEncoder()
for col in ['Gender','Contract_Type','Internet_Service',
            'Tech_Support','Online_Security','Payment_Method','Churn']:
    df_model[col] = le.fit_transform(df_model[col])

# Features and target
X = df_model.drop(columns=['CustomerID','Churn'])
y = df_model['Churn']

# Split train/test
X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=42)

# Train model
model = LogisticRegression(max_iter=1000)
model.fit(X_train, y_train)

# Evaluate
y_pred = model.predict(X_test)
print(classification_report(y_test, y_pred))
print('Accuracy:', round((y_pred == y_test).mean() * 100, 1), '%')"""),
    ("Step 5: Feature importance",
     """import numpy as np

# Get feature importance from model coefficients
feature_names = X.columns
coefficients = model.coef_[0]

importance_df = pd.DataFrame({
    'Feature': feature_names,
    'Importance': np.abs(coefficients)
}).sort_values('Importance', ascending=True)

# Plot
plt.figure(figsize=(8, 6))
plt.barh(importance_df['Feature'], importance_df['Importance'], color='steelblue')
plt.title('Feature Importance — What Drives Churn?', fontsize=13, fontweight='bold')
plt.xlabel('Coefficient Magnitude')
plt.tight_layout()
plt.savefig('feature_importance.png', dpi=150)
plt.show()"""),
]

row_cursor = 3
for title, code in py_blocks:
    ws4.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
    tc = ws4.cell(row_cursor, 1)
    tc.value = title
    tc.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    tc.fill = fill("6C3483")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    tc.border = thin_border
    ws4.row_dimensions[row_cursor].height = 24
    row_cursor += 1

    for line in code.split("\n"):
        ws4.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
        cc = ws4.cell(row_cursor, 1)
        cc.value = line
        cc.font = Font(name="Courier New", size=10, color="1A1A2E")
        cc.fill = fill("F9F4FF")
        cc.alignment = Alignment(horizontal="left", vertical="center")
        cc.border = thin_border
        ws4.row_dimensions[row_cursor].height = 18
        row_cursor += 1

    row_cursor += 1

ws4.column_dimensions["A"].width = 100

# ===================== SHEET 5: POWER BI GUIDE =====================
ws5 = wb.create_sheet("Power BI Guide")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:H1")
ws5["A1"].value = "Power BI Dashboard Guide — Step by Step"
ws5["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
ws5["A1"].fill = fill(DARK_BLUE)
ws5["A1"].alignment = center()
ws5.row_dimensions[1].height = 32

pbi_steps = [
    ("Step 1", "Import Data", "Open Power BI Desktop → Get Data → Excel → select this file → load 'Customer Data' sheet"),
    ("Step 2", "Create Churn Rate Measure", "New Measure → Churn Rate = DIVIDE(CALCULATE(COUNTROWS(customers), customers[Churn]=\"Yes\"), COUNTROWS(customers))"),
    ("Step 3", "KPI Card — Overall Churn Rate", "Insert → Card visual → drag 'Churn Rate' measure → format as percentage"),
    ("Step 4", "Bar Chart — Churn by Contract", "Insert → Clustered Bar Chart → Axis: Contract_Type → Values: Churn Rate measure"),
    ("Step 5", "Donut Chart — Churn Split", "Insert → Donut Chart → Legend: Churn → Values: CustomerID (Count)"),
    ("Step 6", "Line Chart — Tenure vs Churn", "Insert → Line Chart → X-Axis: Tenure_Months → Y-Axis: Churn Rate measure"),
    ("Step 7", "Table — High Risk Customers", "Insert → Table → Add: CustomerID, Contract_Type, Satisfaction_Score, Num_Complaints → Filter Churn = No"),
    ("Step 8", "Slicers for Filtering", "Insert → Slicer → Field: Contract_Type (repeat for Internet_Service and Gender)"),
    ("Step 9", "Color Theme", "View → Themes → choose a professional blue theme → set churn bars to Red, retained to Green"),
    ("Step 10", "Publish", "File → Publish → Share with Power BI Service → copy link for your portfolio/resume"),
]

row_cursor = 3
pbi_headers = ["Step", "Action", "Instructions"]
for c_idx, h in enumerate(pbi_headers, 1):
    cell = ws5.cell(row_cursor, c_idx)
    cell.value = h
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    cell.fill = fill(MED_BLUE)
    cell.alignment = center()
    cell.border = thin_border
ws5.row_dimensions[row_cursor].height = 24
row_cursor += 1

for i, (step, action, instruction) in enumerate(pbi_steps, start=1):
    alt = i % 2 == 0
    bg = "EBF5FB" if alt else WHITE
    for c_idx, val in enumerate([step, action, instruction], 1):
        cell = ws5.cell(row_cursor, c_idx)
        cell.value = val
        cell.fill = fill(bg)
        cell.font = Font(name="Arial", size=10, bold=(c_idx==2), color="000000")
        cell.alignment = Alignment(horizontal="left" if c_idx==3 else "center",
                                   vertical="center", wrap_text=True)
        cell.border = thin_border
    ws5.row_dimensions[row_cursor].height = 36
    row_cursor += 1

ws5.column_dimensions["A"].width = 10
ws5.column_dimensions["B"].width = 22
ws5.column_dimensions["C"].width = 80

# Save
output_path = "Customer_Churn_Project_Shama.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Dataset: {len(df)} rows, Churn Rate: {churn_rate}%")







